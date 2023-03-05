import os

import openpyxl

file_name_template = "statistics/gocosi"


def wirte_to_xls(data):
    file_name = file_name_template + ".xlsx"
    sheet_name = 'latency'
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(file_name)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(['node_num', 'failing_node_num', 'round_num', 'round_time',
                   'latency_avg', 'latency_max', 'latency_min'])
    ws = wb[sheet_name]
    ws.append(data)
    wb.save(file_name)

def wirte_to_xls_neighbours(data):
    file_name = file_name_template + ".xlsx"
    sheet_name = 'latency_neighbours'
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(file_name)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(['node_num', 'failing_node_num', 'neighbours', 'round_num', 'round_time',
                   'latency_avg', 'latency_max', 'latency_min'])
    ws = wb[sheet_name]
    ws.append(data)
    wb.save(file_name)

def wirte_to_xls_p(data):
    file_name = file_name_template + ".xlsx"
    sheet_name = 'latency_parallel'
    if not os.path.exists(file_name):
        wb = openpyxl.Workbook()
    else:
        wb = openpyxl.load_workbook(file_name)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append(['node_num', 'failing_node_num', 'round_num', 'round_time', 'parallel requests',
                 'latency_avg', 'latency_max', 'latency_min'])
    ws = wb[sheet_name]
    ws.append(data)
    wb.save(file_name)
