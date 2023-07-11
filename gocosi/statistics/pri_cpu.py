import numpy as np
import openpyxl
import pandas
import matplotlib.pyplot as plt
from pandas import DataFrame

config = {
    "font.family": 'serif', # 衬线字体
    "font.size": 20, # 相当于小四大小
    "font.serif": ['SimSun'], # 宋体
    "mathtext.fontset": 'stix', # matplotlib渲染数学字体时使用的字体，和Times New Roman差别不大
    'figure.figsize': (10.0, 10.0)
}
plt.rcParams.update(config)
def parse_pb_xls(filename):
    wb = openpyxl.load_workbook(filename)
    data = dict()
    for sheetname in wb.sheetnames:
        if sheetname == "getpub" or sheetname == "resppub":
            continue
        ws = wb[sheetname]
        maxrows = ws.max_row  # 获取最大行
        gas_list = []

        for i in range(maxrows - 1):
            temp_list = []
            for each in ws.iter_cols(min_row=2):
                temp_list.append(each[i].value)
            # print(temp_list)
            gas_list.append(temp_list[3])
        print(gas_list)
        data[sheetname] = np.mean(gas_list)
    return data


def round_time_fig():
    func_list = ['ReadEC', 'EndorsementInc', 'StatusAssetExists', 'SetStatus', 'SelectCP', 'UpdatePropotion']
    # arange返回一个数据，range返回一个list
    # arange 用于创建等差数组

    fig = plt.figure()
    # ax1显示y1  ,ax2显示y2
    # y = [1.50964, 1.16481, 1.249693, 1.257242, 7.344041, 4.464272]
    y = [1.055, 0.844, 0.6755, 0.9025, 1.835, 1.9245]
    for i in range(len(func_list)):
        plt.bar(func_list[i], y[i], color='#1f77b4', width=0.5)
        # plt.text(func_list[i], y[i], "%.f"%y[i], ha='center')

    # plt.xlabel('数据类型', fontdict=font1)
    plt.xticks(rotation=20)
    plt.ylabel('CPU 使用率(%)')
    plt.savefig('./output/pricpu.svg', dpi=300)  # eps文件，用于LaTeX
    plt.show()


if __name__ == '__main__':
    round_time_fig()
    # print(parse_latency_xls("gocosi.xlsx", "latency_round_time"))
    # print(round_time_cpu([5, 10, 50, 100, 500, 800, 1000]))
