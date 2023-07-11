import numpy as np
import openpyxl
import pandas
import matplotlib.pyplot as plt
from pandas import DataFrame

# plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签SimHei
config = {
    "font.family": 'serif', # 衬线字体
    "font.size": 20, # 相当于小四大小
    "font.serif": ['SimSun'], # 宋体
    "mathtext.fontset": 'stix', # matplotlib渲染数学字体时使用的字体，和Times New Roman差别不大
    'figure.figsize': (10.0, 10.0)
}

def parse_pb_xls(filename):
    wb = openpyxl.load_workbook(filename)
    data = dict()
    for sheetname in wb.sheetnames:
        if sheetname == "getpub" or sheetname == "resppub":
            continue
        ws = wb[sheetname]
        maxrows = ws.max_row  # 获取最大行
        gas_list = []

        for i in range(maxrows - 1):
            temp_list = []
            for each in ws.iter_cols(min_row=2):
                temp_list.append(each[i].value)
            # print(temp_list)
            gas_list.append(temp_list[3])
        print(gas_list)
        data[sheetname] = np.mean(gas_list)
    return data


def round_time_fig():
    func_list = ['Use Case 1\nRequest', 'Use Case 1\nResponse', 'Use Case 2\nRequest', 'Use Case 2\nResponse']
    gas_dict = parse_pb_xls("public blockchain.xlsx")
    # arange返回一个数据，range返回一个list
    # arange 用于创建等差数组

    fig = plt.figure()
    font1 = {'family': 'Times New Roman',
             'style': 'normal',
             'size': 10,
             }
    # ax1显示y1  ,ax2显示y2
    y = [0,0,0,0]
    for func in list(gas_dict.keys()):
        index = 0
        if func == 'userreq':
            index = 2
        elif func == 'resp':
            index = 3
        elif func == 'getcategory':
            index = 0
        elif func == 'respcategory':
            index = 1
        else:
            break
        y[index] = gas_dict[func]
    for i in range(len(func_list)):
        plt.bar(func_list[i], y[i], color='#1f77b4', width=0.5, )
        # plt.text(func_list[i], y[i], "%.f"%y[i], ha='center')


    plt.ylabel('Gas Consumption')
    plt.savefig('./output/gas.eps', dpi=300)  # eps文件，用于LaTeX
    plt.show()



if __name__ == '__main__':
    round_time_fig()
    # print(parse_latency_xls("gocosi.xlsx", "latency_round_time"))
    # print(round_time_cpu([5, 10, 50, 100, 500, 800, 1000]))
