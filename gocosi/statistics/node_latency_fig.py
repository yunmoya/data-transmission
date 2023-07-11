import numpy as np
import openpyxl
import pandas
import matplotlib.pyplot as plt
from pandas import DataFrame

# plt.rcParams['font.sans-serif'] = ['SimHei']  # 用来正常显示中文标签SimHei
config = {
    "font.family": 'serif', # 衬线字体
    "font.size": 20, # 相当于小四大小
    "font.serif": ['SimSun'], # 宋体
    "mathtext.fontset": 'stix', # matplotlib渲染数学字体时使用的字体，和Times New Roman差别不大
    'figure.figsize': (10.0, 10.0)
}
plt.rcParams.update(config)

def parse_latency_xls(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    if sheetname not in wb.sheetnames:
        print(f"error sheet name")
        return None
    ws = wb[sheetname]
    maxrows = ws.max_row  # 获取最大行
    data = dict()
    for i in range(maxrows - 1):
        temp_list = []
        for each in ws.iter_cols(min_row=2):
            temp_list.append(each[i].value)
        print(temp_list)
        neighbours = temp_list[2]
        avg_latency = temp_list[5]
        if neighbours not in data:
            data[neighbours] = [avg_latency]
        else:
            data[neighbours].append(avg_latency)
    return data


def node_time_fig():
    node_nums = [4, 10, 20, 30, 40, 50]
    latency_dict = parse_latency_xls("gocosi.xlsx", "latency_neighbours")
    color = {3: 'r', 5: 'b', 10: 'g'}
    line_sty = {3: 'solid', 5: 'dotted', 10: 'dashed'}
    # arange返回一个数据，range返回一个list
    # arange 用于创建等差数组

    fig = plt.figure()
    # ax1显示y1  ,ax2显示y2
    ax1 = fig.subplots()
    for nn in latency_dict:
        y1 = latency_dict[nn]
        ax1.plot(node_nums, y1, f'{color[nn]}', linestyle=f'{line_sty[nn]}', marker='.', label=f'F={nn}')

    ax1.set_xlabel('节点数量')
    ax1.set_ylabel('时延(s)')

    plt.xticks(node_nums)
    plt.legend()
    plt.grid()
    plt.savefig('./output/gocosiF.svg', dpi=300)  # eps文件，用于LaTeX
    plt.show()


if __name__ == '__main__':
    node_time_fig()
    # print(parse_latency_xls("gocosi.xlsx", "latency_round_time"))
    # print(round_time_cpu([5, 10, 50, 100, 500, 800, 1000]))
