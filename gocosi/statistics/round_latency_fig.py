import numpy as np
import openpyxl
import pandas
import matplotlib.pyplot as plt
from pandas import DataFrame

config = {
    "font.family": 'serif', # 衬线字体
    "font.size": 20, # 相当于小四大小
    "font.serif": ['SimSun'], # 宋体
    "mathtext.fontset": 'stix', # matplotlib渲染数学字体时使用的字体，和Times New Roman差别不大
    'figure.figsize': (10.0, 10.0)
}
plt.rcParams.update(config)

def parse_latency_xls(filename, sheetname):
    wb = openpyxl.load_workbook(filename)
    if sheetname not in wb.sheetnames:
        print(f"error sheet name")
        return None
    ws = wb[sheetname]
    maxrows = ws.max_row  # 获取最大行
    data = dict()
    for i in range(maxrows - 1):
        temp_list = []
        for each in ws.iter_cols(min_row=2):
            temp_list.append(each[i].value)
        print(temp_list)
        nn = temp_list[0]
        avg_latency = temp_list[4]
        if nn not in data:
            data[nn] = [avg_latency]
        else:
            data[nn].append(avg_latency)
    return data


def round_time_fig():
    round_times = [50, 100, 150, 200, 250, 300, 350, 400, 450, 500]
    # round_times = [50, 100, 150]
    # latency_dict = parse_latency_xls("gocosilrt.xlsx", "latency_round_time")
    latency_dict = parse_latency_xls("gocosi.xlsx", "latency_round")
    color = {3: 'r', 10: 'b', 20: 'g', 30: 'c'}
    line_sty = {3: 'solid', 10: 'dotted', 20: 'dashed', 30: 'dashdot'}
    # arange返回一个数据，range返回一个list
    # arange 用于创建等差数组

    fig = plt.figure()
    # ax1显示y1  ,ax2显示y2
    ax1 = fig.subplots()
    for nn in latency_dict:
        color_nn = color[nn]
        line_sty_nn = line_sty[nn]
        y1 = latency_dict[nn]
        if nn == 3:
            nn = 4
        ax1.plot(round_times, y1, f'{color_nn}', linestyle=f'{line_sty_nn}', marker='.', label=f'N={nn}')

    ax1.set_xlabel('流言传播周期(ms)')
    ax1.set_ylabel('时延(s)')

    plt.xticks(round_times)
    plt.legend()
    plt.grid()
    plt.savefig('./output/gocosiT.svg', dpi=300)  # eps文件，用于LaTeX
    plt.show()


if __name__ == '__main__':
    round_time_fig()
    # print(parse_latency_xls("gocosi.xlsx", "latency_round_time"))
    # print(round_time_cpu([5, 10, 50, 100, 500, 800, 1000]))
